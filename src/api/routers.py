"""
routers.py
----------
Tous les endpoints REST de l'application.

Alertes   : GET /api/alerts, GET /api/alerts/{id}, PATCH ack/resolve, GET /api/alerts/export
Logs      : GET /api/logs, GET /api/logs/stats, GET /api/logs/mttr
Capteurs  : GET /api/sensors/health
Simulateur: POST /api/simulator/tick, POST /api/simulator/config
Config    : GET/POST /api/config/thresholds, POST /api/config/thresholds/reset
"""

import csv
import io
from fastapi import APIRouter, Depends, HTTPException, Query, Request, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from ..alerts.database import get_db
from ..alerts.service import AlertService
from .auth import get_current_user, can_write
from ..detection.engine import AnomalyDetectionEngine
from ..simulator.generator import IoTSimulator

# Instances partagées (state simple, suffisant pour ce projet)
_simulator = IoTSimulator(anomaly_probability=0.20)
_engine    = AnomalyDetectionEngine(window_size=30)

alerts_router    = APIRouter()
logs_router      = APIRouter()
sensors_router   = APIRouter()
simulator_router = APIRouter()
config_router    = APIRouter()


def require_write(request: Request) -> None:
    """Dépendance FastAPI — lève 403 si l'utilisateur n'a pas le droit d'écriture."""
    if not can_write(request):
        raise HTTPException(status_code=403, detail="Accès refusé : droits insuffisants (viewer)")


# ── /api/alerts ────────────────────────────────────────────────────────────

@alerts_router.get("/export")
def export_alerts_csv(
    sensor:   str | None  = Query(None),
    level:    str | None  = Query(None),
    resolved: bool | None = Query(None),
    db: Session = Depends(get_db),
):
    """Exporte les alertes filtrées en CSV téléchargeable."""
    svc    = AlertService(db)
    alerts = svc.get_all(sensor=sensor, level=level, resolved=resolved, limit=5000)

    output = io.StringIO()
    fieldnames = ["id", "sensor", "value", "unit", "level", "method", "reason",
                  "z_score", "acknowledged", "resolved", "created_at", "resolved_at"]
    writer = csv.DictWriter(output, fieldnames=fieldnames, extrasaction="ignore")
    writer.writeheader()
    for a in alerts:
        writer.writerow(a.to_dict())

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=alertes.csv"},
    )


@alerts_router.get("")
def list_alerts(
    sensor:    str | None  = Query(None),
    level:     str | None  = Query(None),
    resolved:  bool | None = Query(None),
    date_from: str | None  = Query(None, description="ISO date YYYY-MM-DD"),
    date_to:   str | None  = Query(None, description="ISO date YYYY-MM-DD"),
    q:         str | None  = Query(None, description="Recherche dans la raison"),
    limit:  int = Query(50,  ge=1, le=200),
    offset: int = Query(0,   ge=0),
    db: Session = Depends(get_db),
):
    """Liste les alertes avec filtres optionnels."""
    from datetime import datetime, timezone, timedelta
    df = datetime.fromisoformat(date_from).replace(tzinfo=timezone.utc) if date_from else None
    dt = (datetime.fromisoformat(date_to) + timedelta(days=1)).replace(tzinfo=timezone.utc) if date_to else None
    svc    = AlertService(db)
    alerts = svc.get_all(
        sensor=sensor, level=level, resolved=resolved,
        date_from=df, date_to=dt, search=q,
        limit=limit, offset=offset,
    )
    total = svc.count_all(sensor=sensor, level=level, resolved=resolved, date_from=df, date_to=dt, search=q)
    return {"alerts": [a.to_dict() for a in alerts], "count": len(alerts), "total": total}


@alerts_router.get("/open")
def open_alerts_count(db: Session = Depends(get_db)):
    """Nombre d'alertes ouvertes par niveau."""
    return AlertService(db).get_open_count()


@alerts_router.get("/{alert_id}")
def get_alert(alert_id: int, db: Session = Depends(get_db)):
    """Détail d'une alerte."""
    alert = AlertService(db).get_by_id(alert_id)
    if not alert:
        raise HTTPException(status_code=404, detail="Alerte non trouvée")
    return alert.to_dict()


@alerts_router.patch("/{alert_id}/acknowledge")
def acknowledge_alert(
    alert_id: int,
    request: Request,
    reason: str = Query("", description="Raison de l'acquittement"),
    db: Session = Depends(get_db),
    _: None = Depends(require_write),
):
    alert = AlertService(db, user=get_current_user(request) or "système").acknowledge(alert_id, reason=reason)
    if not alert:
        raise HTTPException(status_code=404, detail="Alerte non trouvée")
    return {"message": "Alerte acquittée", "alert": alert.to_dict()}


@alerts_router.patch("/{alert_id}/resolve")
def resolve_alert(alert_id: int, request: Request, db: Session = Depends(get_db), _: None = Depends(require_write)):
    alert = AlertService(db, user=get_current_user(request) or "système").resolve(alert_id)
    if not alert:
        raise HTTPException(status_code=404, detail="Alerte non trouvée")
    return {"message": "Alerte résolue", "alert": alert.to_dict()}


@alerts_router.patch("/{alert_id}/snooze")
def snooze_alert(
    alert_id: int,
    request: Request,
    minutes: int = Query(30, ge=1, le=1440, description="Durée du snooze en minutes"),
    db: Session = Depends(get_db),
    _: None = Depends(require_write),
):
    """Met en sourdine une alerte pendant N minutes."""
    alert = AlertService(db, user=get_current_user(request) or "système").snooze(alert_id, minutes=minutes)
    if not alert:
        raise HTTPException(status_code=404, detail="Alerte non trouvée")
    return {"message": f"Alerte snoozée {minutes} min", "alert": alert.to_dict()}


@alerts_router.patch("/{alert_id}/unsnooze")
def unsnooze_alert(alert_id: int, request: Request, db: Session = Depends(get_db), _: None = Depends(require_write)):
    """Annule le snooze d'une alerte."""
    alert = AlertService(db, user=get_current_user(request) or "système").unsnooze(alert_id)
    if not alert:
        raise HTTPException(status_code=404, detail="Alerte non trouvée")
    return {"message": "Snooze annulé", "alert": alert.to_dict()}


@alerts_router.patch("/{alert_id}/notes")
def update_note(alert_id: int, request: Request, note: str = "", db: Session = Depends(get_db)):
    """Ajoute ou met à jour la note opérateur d'une alerte."""
    alert = AlertService(db, user=get_current_user(request) or "système").add_note(alert_id, note)
    if not alert:
        raise HTTPException(status_code=404, detail="Alerte non trouvée")
    return {"message": "Note enregistrée", "alert": alert.to_dict()}


@alerts_router.patch("/{alert_id}/tags")
def update_tags(alert_id: int, request: Request, tags: str = Query("", description="Tags séparés par virgules"), db: Session = Depends(get_db)):
    """Définit les tags d'une alerte."""
    tag_list = [t for t in tags.split(",") if t.strip()]
    alert = AlertService(db, user=get_current_user(request) or "système").set_tags(alert_id, tag_list)
    if not alert:
        raise HTTPException(status_code=404, detail="Alerte non trouvée")
    return {"message": "Tags enregistrés", "alert": alert.to_dict()}


@alerts_router.delete("/{alert_id}")
def delete_alert(alert_id: int, request: Request, db: Session = Depends(get_db)):
    """Supprime définitivement une alerte."""
    deleted = AlertService(db, user=get_current_user(request) or "système").delete(alert_id)
    if not deleted:
        raise HTTPException(status_code=404, detail="Alerte non trouvée")
    return {"message": "Alerte supprimée"}


@alerts_router.delete("")
def delete_alerts_bulk(
    request: Request,
    ids: str = Query(..., description="IDs séparés par virgules, ex: 1,2,3"),
    db: Session = Depends(get_db),
):
    """Supprime plusieurs alertes en une seule requête."""
    try:
        id_list = [int(i.strip()) for i in ids.split(",") if i.strip()]
    except ValueError:
        raise HTTPException(status_code=400, detail="IDs invalides")
    count = AlertService(db, user=get_current_user(request) or "système").delete_bulk(id_list)
    return {"message": f"{count} alerte(s) supprimée(s)", "count": count}


@alerts_router.patch("/acknowledge-all")
def acknowledge_all(request: Request, db: Session = Depends(get_db)):
    count = AlertService(db, user=get_current_user(request) or "système").acknowledge_all()
    return {"message": f"{count} alerte(s) acquittée(s)"}


@alerts_router.get("/export-json")
def export_alerts_json(
    sensor:   str | None  = Query(None),
    level:    str | None  = Query(None),
    resolved: bool | None = Query(None),
    db: Session = Depends(get_db),
):
    """Exporte les alertes filtrées en JSON téléchargeable."""
    import json as _json
    svc    = AlertService(db)
    alerts = svc.get_all(sensor=sensor, level=level, resolved=resolved, limit=5000)
    payload = _json.dumps([a.to_dict() for a in alerts], indent=2, ensure_ascii=False)
    return StreamingResponse(
        iter([payload]),
        media_type="application/json",
        headers={"Content-Disposition": "attachment; filename=alertes.json"},
    )


@alerts_router.post("/archive")
def archive_alerts(request: Request, days: int = Query(0, ge=0, le=365), db: Session = Depends(get_db)):
    """Archive les alertes résolues de plus de N jours (0 = toutes les résolues)."""
    count = AlertService(db, user=get_current_user(request) or "système").archive_old_alerts(days=days)
    return {"message": f"{count} alerte(s) archivée(s)", "count": count}


@alerts_router.post("/archived/{archived_id}/unarchive")
def unarchive_alert(archived_id: int, request: Request, db: Session = Depends(get_db)):
    """Remet une alerte archivée dans la table alerts."""
    alert = AlertService(db, user=get_current_user(request) or "système").unarchive(archived_id)
    if not alert:
        raise HTTPException(status_code=404, detail="Alerte archivée non trouvée")
    return {"message": "Alerte restaurée", "alert": alert.to_dict()}


@alerts_router.get("/{alert_id}/comments")
def get_comments(alert_id: int, db: Session = Depends(get_db)):
    """Retourne les commentaires d'une alerte."""
    comments = AlertService(db).get_comments(alert_id)
    return {"comments": [c.to_dict() for c in comments]}


@alerts_router.post("/{alert_id}/comments")
def add_comment(
    alert_id: int,
    request: Request,
    content: str = Query(..., description="Contenu du commentaire"),
    db: Session = Depends(get_db),
):
    """Ajoute un commentaire à une alerte."""
    comment = AlertService(db, user=get_current_user(request) or "système").add_comment(alert_id, content)
    if not comment:
        raise HTTPException(status_code=404, detail="Alerte non trouvée")
    return {"message": "Commentaire ajouté", "comment": comment.to_dict()}


@alerts_router.delete("/{alert_id}/comments/{comment_id}")
def delete_comment(alert_id: int, comment_id: int, db: Session = Depends(get_db)):
    """Supprime un commentaire."""
    deleted = AlertService(db).delete_comment(comment_id)
    if not deleted:
        raise HTTPException(status_code=404, detail="Commentaire non trouvé")
    return {"message": "Commentaire supprimé"}


# ── /api/ingest — données réelles depuis capteurs externes ────
ingest_router = APIRouter()

def _validate_device_key(request: Request, db: Session) -> bool:
    """Valide le header X-Device-Key si présent. Retourne True si valide ou aucune clé configurée."""
    import hashlib
    from ..alerts.models import DeviceKey
    from datetime import datetime, timezone
    header_key = request.headers.get("X-Device-Key", "")
    if not header_key:
        return True  # Pas de clé = accès libre (compatible rétroactivement)
    key_hash = hashlib.sha256(header_key.encode()).hexdigest()
    now = datetime.now(timezone.utc)
    k = db.query(DeviceKey).filter(
        DeviceKey.key_hash == key_hash,
        DeviceKey.active == True,
    ).first()
    if not k:
        return False
    if k.expires_at and k.expires_at < now:
        return False
    k.last_used_at = now
    db.commit()
    return True


@ingest_router.post("")
def ingest_reading(
    request: Request,
    sensor: str  = Query(..., description="Nom du capteur : temperature | turbidity | ph"),
    value:  float = Query(..., description="Valeur mesurée"),
    unit:   str  = Query("",  description="Unité de mesure"),
    api_key: str = Query("", description="Clé API optionnelle (legacy)"),
    db: Session = Depends(get_db),
):
    """
    Endpoint pour soumettre une lecture depuis un vrai capteur IoT.
    Exemple : POST /api/ingest?sensor=temperature&value=28.5&unit=°C
    Header optionnel : X-Device-Key: <clé>
    """
    if not _validate_device_key(request, db):
        raise HTTPException(status_code=401, detail="Clé API appareil invalide ou expirée")
    VALID = {"temperature", "turbidity", "ph"}
    if sensor not in VALID:
        raise HTTPException(400, f"Capteur invalide — valeurs acceptées : {', '.join(VALID)}")

    from ..simulator.generator import SensorReading
    reading = SensorReading(sensor=sensor, value=value, scenario="external")

    result  = _engine.analyze(reading)
    svc     = AlertService(db, user="capteur_externe")
    alert   = svc.process(result, reading)

    return {
        "sensor":  sensor,
        "value":   value,
        "level":   result.level.value,
        "reason":  result.reason,
        "alert":   alert.to_dict() if alert else None,
    }


@ingest_router.post("/upload")
async def ingest_upload(
    file: UploadFile,
    db:   Session = Depends(get_db),
):
    """
    Import de lectures IoT depuis un fichier CSV ou JSON.
    CSV attendu : sensor,value,unit  (une ligne d'en-tête)
    JSON attendu: [{"sensor":"temperature","value":22.5,"unit":"°C"}, ...]
    """
    import csv, io, json as _json
    VALID = {"temperature", "turbidity", "ph"}

    content = await file.read()
    rows = []

    if file.filename.endswith(".json"):
        try:
            data = _json.loads(content.decode("utf-8", errors="replace"))
            rows = data if isinstance(data, list) else []
        except Exception:
            raise HTTPException(400, "Fichier JSON invalide")
    else:
        # CSV
        try:
            reader = csv.DictReader(io.StringIO(content.decode("utf-8", errors="replace")))
            for r in reader:
                rows.append({"sensor": r.get("sensor","").strip(),
                             "value":  r.get("value","0").strip(),
                             "unit":   r.get("unit","").strip()})
        except Exception:
            raise HTTPException(400, "Fichier CSV invalide")

    from ..simulator.generator import SensorReading
    svc = AlertService(db, user="import_fichier")
    created, skipped, alerts_n = 0, 0, 0

    for row in rows:
        sensor = str(row.get("sensor","")).strip()
        if sensor not in VALID:
            skipped += 1; continue
        try:
            value = float(row.get("value", 0))
        except (ValueError, TypeError):
            skipped += 1; continue
        unit = str(row.get("unit","")).strip() or "?"

        reading = SensorReading(sensor=sensor, value=value, scenario="import")
        result  = _engine.analyze(reading)
        alert   = svc.process(result, reading)
        created += 1
        if alert: alerts_n += 1

    return {"imported": created, "skipped": skipped, "alerts_created": alerts_n}


# ── /api/logs ─────────────────────────────────────────────────────────────

@logs_router.get("/stats")
def get_stats(db: Session = Depends(get_db)):
    """Statistiques globales des alertes."""
    return AlertService(db).get_stats()


@logs_router.get("/mttr")
def get_mttr(db: Session = Depends(get_db)):
    """Mean Time To Resolve par capteur (secondes)."""
    return AlertService(db).get_mttr()


@logs_router.get("/heatmap")
def get_heatmap(days: int = Query(7, ge=1, le=30), db: Session = Depends(get_db)):
    """Heatmap des anomalies par heure/jour sur les N derniers jours."""
    return AlertService(db).get_heatmap(days=days)


@logs_router.get("/correlation")
def get_correlation(days: int = Query(1, ge=1, le=30), db: Session = Depends(get_db)):
    """Paires de valeurs capteurs pour graphiques de corrélation."""
    return AlertService(db).get_correlation(days=days)


@logs_router.get("/open-duration")
def get_open_duration(days: int = Query(7, ge=1, le=30), db: Session = Depends(get_db)):
    """Durée moyenne des alertes non résolues par niveau."""
    return AlertService(db).get_open_duration(days=days)


@logs_router.get("/trend")
def get_trend(
    days:   int = Query(30, ge=1, le=90),
    sensor: str | None = Query(None),
    db: Session = Depends(get_db),
):
    """Agrégation journalière des lectures par capteur sur N jours."""
    return AlertService(db).get_trend(days=days, sensor=sensor)


@logs_router.get("/compare")
def get_comparison(
    period: str = Query("week", description="week | month"),
    db: Session = Depends(get_db),
):
    """Statistiques période courante vs période précédente."""
    if period not in ("week", "month"):
        from fastapi import HTTPException
        raise HTTPException(400, "period doit être 'week' ou 'month'")
    return AlertService(db).get_comparison(period=period)


@logs_router.get("/composite")
def get_composite_alerts(db: Session = Depends(get_db)):
    """Liste des alertes composites (corrélation multi-capteurs)."""
    svc = AlertService(db)
    alerts = svc.get_all(sensor="composite", limit=50)
    return {"alerts": [a.to_dict() for a in alerts], "count": len(alerts)}


@logs_router.get("/audit")
def get_audit(
    alert_id:  int | None  = Query(None),
    action:    str | None  = Query(None, description="Filtrer par action"),
    date_from: str | None  = Query(None, description="ISO date YYYY-MM-DD"),
    date_to:   str | None  = Query(None, description="ISO date YYYY-MM-DD"),
    limit:  int = Query(50, ge=1, le=200),
    offset: int = Query(0, ge=0),
    db: Session = Depends(get_db),
):
    """Journal d'audit des actions opérateur."""
    from datetime import datetime, timezone, timedelta
    df = datetime.fromisoformat(date_from).replace(tzinfo=timezone.utc) if date_from else None
    dt = (datetime.fromisoformat(date_to) + timedelta(days=1)).replace(tzinfo=timezone.utc) if date_to else None
    svc   = AlertService(db)
    logs  = svc.get_audit_log(alert_id=alert_id, action=action, date_from=df, date_to=dt, limit=limit, offset=offset)
    total = svc.count_audit_log(alert_id=alert_id, action=action, date_from=df, date_to=dt)
    return {"logs": [l.to_dict() for l in logs], "total": total}


@logs_router.post("/compare-two")
async def compare_two_files(file1: UploadFile, file2: UploadFile):
    """Compare deux fichiers CSV/Excel exportés et retourne leurs statistiques côte à côte."""
    import csv as _csv, io as _io

    async def parse_file(f: UploadFile) -> tuple[list[dict], str]:
        fname   = f.filename or ""
        content = await f.read()
        if fname.endswith(".pdf"):
            raise HTTPException(400, f"'{fname}' : PDF non supporté. Utilisez CSV ou Excel.")
        rows = []
        if fname.endswith(".xlsx"):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(_io.BytesIO(content), data_only=True)
                ws = wb.active
                headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    rows.append(dict(zip(headers, row)))
            except HTTPException:
                raise
            except Exception as e:
                raise HTTPException(400, f"'{fname}' : Excel invalide — {e}")
        else:
            try:
                text  = content.decode("utf-8", errors="replace")
                lines = [l for l in text.splitlines() if not l.startswith("#")]
                rows  = list(_csv.DictReader(lines))
            except Exception as e:
                raise HTTPException(400, f"'{fname}' : CSV invalide — {e}")
        return rows, fname

    def compute_stats(rows: list[dict]) -> dict:
        sensors: dict[str, dict] = {}
        skipped = 0
        for row in rows:
            sensor = str(row.get("sensor") or row.get("Capteur") or "").strip()
            level  = str(row.get("level")  or row.get("Niveau")  or "NORMAL").strip().upper()
            if not sensor:
                skipped += 1
                continue
            if sensor not in sensors:
                sensors[sensor] = {"total": 0, "clean": 0, "aberrant": 0}
            sensors[sensor]["total"] += 1
            if level == "NORMAL":
                sensors[sensor]["clean"] += 1
            else:
                sensors[sensor]["aberrant"] += 1
        stats = []
        for s, d in sorted(sensors.items()):
            pct = round(d["aberrant"] / d["total"] * 100, 1) if d["total"] else 0
            stats.append({"sensor": s, "total": d["total"], "clean": d["clean"],
                          "aberrant": d["aberrant"], "pct": pct})
        total_all    = sum(s["total"]    for s in stats)
        clean_all    = sum(s["clean"]    for s in stats)
        aberrant_all = sum(s["aberrant"] for s in stats)
        pct_all      = round(aberrant_all / total_all * 100, 1) if total_all else 0
        return {"stats": stats, "total_all": total_all, "clean_all": clean_all,
                "aberrant_all": aberrant_all, "pct_all": pct_all, "skipped": skipped}

    rows1, fname1 = await parse_file(file1)
    rows2, fname2 = await parse_file(file2)

    if not rows1: raise HTTPException(400, f"'{fname1}' est vide.")
    if not rows2: raise HTTPException(400, f"'{fname2}' est vide.")

    s1 = compute_stats(rows1)
    s2 = compute_stats(rows2)

    # Différences ligne à ligne par capteur
    all_sensors = sorted(set(
        [s["sensor"] for s in s1["stats"]] + [s["sensor"] for s in s2["stats"]]
    ))
    def find(stats_list, sensor):
        return next((s for s in stats_list if s["sensor"] == sensor),
                    {"total": 0, "clean": 0, "aberrant": 0, "pct": 0})

    diff = []
    for sensor in all_sensors:
        a = find(s1["stats"], sensor)
        b = find(s2["stats"], sensor)
        diff.append({
            "sensor":        sensor,
            "delta_total":   b["total"]    - a["total"],
            "delta_clean":   b["clean"]    - a["clean"],
            "delta_aberrant":b["aberrant"] - a["aberrant"],
            "delta_pct":     round(b["pct"] - a["pct"], 1),
        })

    return {
        "file1": {"filename": fname1, **s1},
        "file2": {"filename": fname2, **s2},
        "diff":  diff,
        "delta_total":    s2["total_all"]    - s1["total_all"],
        "delta_clean":    s2["clean_all"]    - s1["clean_all"],
        "delta_aberrant": s2["aberrant_all"] - s1["aberrant_all"],
        "delta_pct":      round(s2["pct_all"] - s1["pct_all"], 1),
    }


@logs_router.post("/compare-upload")
async def compare_upload(file: UploadFile):
    """
    Analyse un fichier CSV ou Excel exporté depuis l'application
    et retourne les statistiques de comparaison complet/nettoyé.
    Colonnes attendues : sensor, value, unit, level, scenario, created_at
    """
    import csv as _csv, io as _io, json as _json

    fname = file.filename or ""
    content = await file.read()

    if fname.endswith(".pdf"):
        raise HTTPException(400, "Le format PDF ne peut pas être analysé comme données. Utilisez CSV ou Excel.")

    rows = []
    if fname.endswith(".xlsx"):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(_io.BytesIO(content), data_only=True)
            ws = wb.active
            headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append(dict(zip(headers, row)))
        except Exception as e:
            raise HTTPException(400, f"Fichier Excel invalide : {e}")

    elif fname.endswith(".json"):
        try:
            rows = _json.loads(content.decode("utf-8", errors="replace"))
            if not isinstance(rows, list):
                raise HTTPException(400, "JSON invalide : tableau attendu")
        except HTTPException:
            raise
        except Exception as e:
            raise HTTPException(400, f"Fichier JSON invalide : {e}")

    else:
        # CSV — ignorer les lignes commentaires (#)
        try:
            text  = content.decode("utf-8", errors="replace")
            lines = [l for l in text.splitlines() if not l.startswith("#")]
            rows  = list(_csv.DictReader(lines))
        except Exception as e:
            raise HTTPException(400, f"Fichier CSV invalide : {e}")

    if not rows:
        raise HTTPException(400, "Le fichier est vide ou illisible.")

    # Calcul des stats par capteur
    sensors: dict[str, dict] = {}
    skipped = 0
    for row in rows:
        sensor = str(row.get("sensor") or row.get("Capteur") or "").strip()
        level  = str(row.get("level")  or row.get("Niveau")  or "NORMAL").strip().upper()
        if not sensor:
            skipped += 1
            continue
        if sensor not in sensors:
            sensors[sensor] = {"total": 0, "clean": 0, "aberrant": 0}
        sensors[sensor]["total"] += 1
        if level == "NORMAL":
            sensors[sensor]["clean"] += 1
        else:
            sensors[sensor]["aberrant"] += 1

    stats = []
    for s, d in sorted(sensors.items()):
        pct = round(d["aberrant"] / d["total"] * 100, 1) if d["total"] else 0
        stats.append({"sensor": s, "total": d["total"], "clean": d["clean"],
                      "aberrant": d["aberrant"], "pct": pct})

    total_all    = sum(s["total"]    for s in stats)
    clean_all    = sum(s["clean"]    for s in stats)
    aberrant_all = sum(s["aberrant"] for s in stats)
    pct_all      = round(aberrant_all / total_all * 100, 1) if total_all else 0

    return {
        "filename":     fname,
        "stats":        stats,
        "total_all":    total_all,
        "clean_all":    clean_all,
        "aberrant_all": aberrant_all,
        "pct_all":      pct_all,
        "skipped":      skipped,
    }


@logs_router.get("/export")
def export_logs_csv(
    sensor:           str | None = Query(None),
    date_from:        str | None = Query(None, description="ISO date YYYY-MM-DD"),
    date_to:          str | None = Query(None, description="ISO date YYYY-MM-DD"),
    include_outliers: bool       = Query(True,  description="False = exclure WARNING/CRITICAL"),
    db: Session = Depends(get_db),
):
    """Exporte les lectures capteurs en CSV (avec ou sans valeurs aberrantes)."""
    from datetime import datetime, timezone, timedelta
    df = datetime.fromisoformat(date_from).replace(tzinfo=timezone.utc) if date_from else None
    dt = (datetime.fromisoformat(date_to) + timedelta(days=1)).replace(tzinfo=timezone.utc) if date_to else None
    svc = AlertService(db)

    total_count    = svc.count_logs(sensor=sensor, date_from=df, date_to=dt, exclude_outliers=False)
    clean_count    = svc.count_logs(sensor=sensor, date_from=df, date_to=dt, exclude_outliers=True)
    aberrant_count = total_count - clean_count

    logs = svc.get_logs(
        sensor=sensor, date_from=df, date_to=dt,
        limit=10000, offset=0,
        exclude_outliers=not include_outliers,
    )

    output = io.StringIO()
    # -- Bloc de description --
    if include_outliers:
        output.write(f"# Export : Données complètes (valeurs aberrantes INCLUSES)\n")
        output.write(f"# Valeurs aberrantes présentes : {aberrant_count} lecture(s) sur {total_count}\n")
    else:
        output.write(f"# Export : Données nettoyées (valeurs aberrantes SUPPRIMÉES)\n")
        output.write(f"# Lignes retirées : {aberrant_count} lecture(s) | Conservées : {clean_count} sur {total_count}\n")
    output.write(f"# Généré le : {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
    if sensor:
        output.write(f"# Capteur filtré : {sensor}\n")
    output.write("#\n")

    fieldnames = ["id", "sensor", "value", "unit", "level", "scenario", "created_at"]
    writer     = csv.DictWriter(output, fieldnames=fieldnames, extrasaction="ignore")
    writer.writeheader()
    for l in logs:
        writer.writerow(l.to_dict())
    output.seek(0)
    suffix = "brut" if include_outliers else "nettoye"
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=lectures_{suffix}.csv"},
    )


@logs_router.get("")
def list_logs(
    sensor:           str | None = Query(None),
    date_from:        str | None = Query(None, description="ISO date YYYY-MM-DD"),
    date_to:          str | None = Query(None, description="ISO date YYYY-MM-DD"),
    limit:            int        = Query(100, ge=1, le=10000),
    offset:           int        = Query(0,   ge=0),
    exclude_outliers: bool       = Query(False, description="True = exclure WARNING/CRITICAL"),
    db: Session = Depends(get_db),
):
    """Historique des lectures capteurs."""
    from datetime import datetime, timezone, timedelta
    df = datetime.fromisoformat(date_from).replace(tzinfo=timezone.utc) if date_from else None
    dt = (datetime.fromisoformat(date_to) + timedelta(days=1)).replace(tzinfo=timezone.utc) if date_to else None
    svc   = AlertService(db)
    logs  = svc.get_logs(sensor=sensor, date_from=df, date_to=dt, limit=limit, offset=offset, exclude_outliers=exclude_outliers)
    total = svc.count_logs(sensor=sensor, date_from=df, date_to=dt, exclude_outliers=exclude_outliers)
    return {"logs": [l.to_dict() for l in logs], "count": len(logs), "total": total}


# ── /api/sensors ──────────────────────────────────────────────────────────

@sensors_router.get("/health")
def sensor_health(db: Session = Depends(get_db)):
    """Dernier état connu (niveau + valeur) pour chaque capteur."""
    return AlertService(db).get_sensor_health()


# ── /api/simulator ────────────────────────────────────────────────────────

@simulator_router.post("/tick")
def simulator_tick(db: Session = Depends(get_db)):
    """Déclenche manuellement un cycle complet (3 capteurs)."""
    readings = _simulator.read_all()
    results  = _engine.analyze_batch(readings)
    svc      = AlertService(db)
    alerts   = svc.process_batch(results, readings)
    return {
        "readings":       [r.to_dict() for r in readings],
        "results":        [r.to_dict() for r in results],
        "alerts_created": [a.to_dict() for a in alerts],
    }


@simulator_router.post("/tick/{sensor}")
def simulator_tick_sensor(sensor: str, db: Session = Depends(get_db)):
    """Déclenche un cycle pour un seul capteur."""
    try:
        reading = _simulator.read_sensor(sensor)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    result = _engine.analyze(reading)
    svc    = AlertService(db)
    alert  = svc.process(result, reading)
    return {
        "reading": reading.to_dict(),
        "result":  result.to_dict(),
        "alert":   alert.to_dict() if alert else None,
    }


@simulator_router.get("/status")
def simulator_status():
    """Retourne l'état courant du simulateur (running ou non)."""
    from .main import get_scheduler
    s = get_scheduler()
    return {"running": s.is_running if s else False}


@simulator_router.post("/start")
def simulator_start():
    """Démarre le simulateur si arrêté."""
    from .main import get_scheduler
    s = get_scheduler()
    if not s:
        raise HTTPException(status_code=503, detail="Scheduler non initialisé")
    if s.is_running:
        return {"running": True, "message": "Déjà en cours"}
    s.start()
    return {"running": True, "message": "Simulateur démarré"}


@simulator_router.post("/stop")
def simulator_stop():
    """Arrête le simulateur si en cours."""
    from .main import get_scheduler
    s = get_scheduler()
    if not s:
        raise HTTPException(status_code=503, detail="Scheduler non initialisé")
    if not s.is_running:
        return {"running": False, "message": "Déjà arrêté"}
    s.stop()
    return {"running": False, "message": "Simulateur arrêté"}


@simulator_router.post("/config")
def set_simulator_config(
    probability: float = Query(0.20, ge=0.0, le=1.0, description="Probabilité d'anomalie (0–1)"),
):
    """Modifie la probabilité d'anomalie du simulateur (en temps réel)."""
    from .stream import set_simulator_probability
    _simulator.anomaly_probability = probability
    set_simulator_probability(probability)   # synchronise aussi le simulateur SSE
    return {"anomaly_probability": probability}


# ── /api/config ───────────────────────────────────────────────────────────

@config_router.get("/thresholds")
def get_thresholds():
    """Retourne la configuration complète des seuils (base + overrides)."""
    from ..alerts.threshold_config import get_all
    return get_all()


@config_router.post("/thresholds")
def update_threshold(
    sensor:  str   = Query(..., description="temperature | turbidity | ph"),
    zone:    str   = Query(..., description="normal | warning | critical"),
    min_val: float | None = Query(None, alias="min"),
    max_val: float | None = Query(None, alias="max"),
):
    """Met à jour un seuil pour un capteur et une zone donnés."""
    from ..alerts.threshold_config import update_sensor
    if sensor not in ("temperature", "turbidity", "ph"):
        raise HTTPException(400, "Capteur invalide")
    if zone not in ("normal", "warning", "critical"):
        raise HTTPException(400, "Zone invalide")
    return update_sensor(sensor, zone, min_val, max_val)


@config_router.post("/thresholds/reset")
def reset_thresholds():
    """Réinitialise tous les seuils aux valeurs par défaut."""
    from ..alerts.threshold_config import reset_all
    reset_all()
    return {"message": "Seuils réinitialisés aux valeurs par défaut"}


@config_router.get("/thresholds/adaptive")
def get_adaptive_thresholds(
    days: int = Query(7, ge=1, le=30),
    db: Session = Depends(get_db),
):
    """Calcule des seuils adaptatifs basés sur l'historique des N derniers jours."""
    return AlertService(db).compute_adaptive_thresholds(days=days)


@config_router.post("/thresholds/apply-adaptive")
def apply_adaptive_thresholds(
    days:   int = Query(7, ge=1, le=30),
    sensor: str = Query(..., description="temperature | turbidity | ph"),
    db: Session = Depends(get_db),
):
    """Applique les seuils adaptatifs calculés pour un capteur donné."""
    from ..alerts.threshold_config import update_sensor
    if sensor not in ("temperature", "turbidity", "ph"):
        raise HTTPException(400, "Capteur invalide")
    suggestions = AlertService(db).compute_adaptive_thresholds(days=days)
    s = suggestions.get(sensor, {})
    if "error" in s:
        raise HTTPException(400, s["error"])
    sg = s["suggested"]
    for zone in ("normal", "warning", "critical"):
        update_sensor(sensor, zone, sg[zone]["min"], sg[zone]["max"])
    return {"message": f"Seuils adaptatifs appliqués pour {sensor}", "applied": sg}


@config_router.get("/thresholds/history")
def get_threshold_history(limit: int = Query(30, ge=1, le=100)):
    """Retourne l'historique des modifications de seuils."""
    from ..alerts.threshold_config import get_history
    return {"history": get_history(limit=limit)}


@config_router.get("/webhook")
def get_webhook():
    """Retourne la configuration webhook actuelle."""
    from ..alerts.notifier import get_config
    return get_config()


@config_router.post("/webhook")
def set_webhook(
    url:    str  = Query("", description="URL du webhook"),
    active: bool = Query(False),
    format: str  = Query("generic", description="generic | slack | discord"),
):
    """Configure le webhook de notification (alertes CRITICAL)."""
    from ..alerts.notifier import set_config
    return set_config(url=url, active=active, fmt=format)


@config_router.get("/maintenance")
def get_maintenance_windows(db: Session = Depends(get_db)):
    """Retourne toutes les fenêtres de maintenance."""
    windows = AlertService(db).get_maintenance_windows()
    return {"windows": [w.to_dict() for w in windows]}


@config_router.post("/maintenance")
def create_maintenance_window(
    request:  Request,
    sensor:   str | None = Query(None, description="Capteur ciblé (None = tous)"),
    start_dt: str = Query(..., description="Début ISO datetime"),
    end_dt:   str = Query(..., description="Fin ISO datetime"),
    reason:   str = Query("", description="Raison de la maintenance"),
    db: Session = Depends(get_db),
):
    """Crée une fenêtre de maintenance."""
    from datetime import datetime, timezone
    try:
        sd = datetime.fromisoformat(start_dt).replace(tzinfo=timezone.utc)
        ed = datetime.fromisoformat(end_dt).replace(tzinfo=timezone.utc)
    except ValueError:
        raise HTTPException(400, "Format datetime invalide (utiliser ISO 8601)")
    if ed <= sd:
        raise HTTPException(400, "La date de fin doit être postérieure à la date de début")
    svc = AlertService(db, user=get_current_user(request) or "système")
    mw  = svc.create_maintenance_window(sensor=sensor, start_dt=sd, end_dt=ed, reason=reason)
    return {"message": "Fenêtre de maintenance créée", "window": mw.to_dict()}


@config_router.delete("/maintenance/{mw_id}")
def delete_maintenance_window(mw_id: int, request: Request, db: Session = Depends(get_db)):
    """Supprime une fenêtre de maintenance."""
    deleted = AlertService(db, user=get_current_user(request) or "système").delete_maintenance_window(mw_id)
    if not deleted:
        raise HTTPException(404, "Fenêtre de maintenance non trouvée")
    return {"message": "Fenêtre supprimée"}


@config_router.post("/report/send-email")
def send_report_email(
    days: int = Query(1, ge=1, le=30),
    db: Session = Depends(get_db),
):
    """Envoie le rapport par email SMTP."""
    from ..alerts.email_notifier import send_report
    svc  = AlertService(db)
    data = svc.get_report_data(days=days)
    sent = send_report(data, days)
    if not sent:
        raise HTTPException(400, "Email non configuré ou inactif — vérifier la configuration SMTP")
    return {"message": f"Rapport {days}j envoyé par email"}


@config_router.get("/report-schedule")
def get_report_schedule():
    """Retourne la configuration du rapport automatique."""
    from ..alerts.report_scheduler import load_report_schedule
    return load_report_schedule()


@config_router.post("/report-schedule")
def set_report_schedule(
    active:      bool = Query(False),
    frequency:   str  = Query("daily", description="daily | weekly"),
    hour:        int  = Query(8, ge=0, le=23),
    days_period: int  = Query(1, description="1 | 7 | 30"),
):
    """Configure l'envoi automatique du rapport par email."""
    from ..alerts.report_scheduler import save_report_schedule
    if frequency not in ("daily", "weekly"):
        raise HTTPException(400, "frequency doit être 'daily' ou 'weekly'")
    return save_report_schedule(active=active, frequency=frequency, hour=hour, days_period=days_period)


@config_router.get("/email")
def get_email_config():
    """Retourne la configuration email SMTP (sans mot de passe)."""
    from ..alerts.email_notifier import get_config
    return get_config()


@config_router.get("/auth")
def get_auth_config_endpoint():
    """Retourne la configuration d'authentification."""
    from .auth import get_auth_config
    return get_auth_config()


@config_router.post("/auth")
def set_auth_config_endpoint(
    enabled:  bool = Query(False),
    username: str  = Query("admin"),
    password: str  = Query(""),
):
    """Configure l'authentification (activer/désactiver + changer le mot de passe)."""
    from .auth import update_auth_config
    return update_auth_config(enabled=enabled, username=username, password=password or None)


@config_router.get("/auth/users")
def list_users(request: Request):
    """Liste les utilisateurs (admin uniquement)."""
    from .auth import get_auth_config, is_admin
    if not is_admin(request):
        raise HTTPException(403, "Accès réservé aux administrateurs")
    return get_auth_config()


@config_router.post("/auth/users")
def create_user(
    request:      Request,
    username:     str = Query(...),
    password:     str = Query(...),
    role:         str = Query("operator", description="admin | operator | viewer"),
    display_name: str = Query("", description="Nom affiché"),
    email:        str = Query("", description="Adresse email"),
    phone:        str = Query("", description="Téléphone"),
):
    """Crée un utilisateur avec un rôle et des données personnelles."""
    from .auth import add_user, is_admin
    if not is_admin(request):
        raise HTTPException(403, "Accès réservé aux administrateurs")
    if role not in ("admin", "operator", "viewer"):
        raise HTTPException(400, "Rôle invalide")
    return add_user(username, password, role, display_name=display_name, email=email, phone=phone)


@config_router.get("/auth/profile")
def get_my_profile(request: Request):
    """Retourne le profil de l'utilisateur connecté."""
    from .auth import get_user_profile, get_current_user
    username = get_current_user(request) or "guest"
    return get_user_profile(username)


@config_router.post("/auth/profile")
def update_my_profile(
    request:      Request,
    display_name: str | None = Query(None),
    email:        str | None = Query(None),
    phone:        str | None = Query(None),
    password:     str | None = Query(None),
):
    """Met à jour le profil de l'utilisateur connecté."""
    from .auth import update_user_profile, get_current_user
    username = get_current_user(request) or "guest"
    return update_user_profile(username, display_name=display_name, email=email, phone=phone, password=password or None)


@config_router.get("/auth/users/{username}/profile")
def get_user_profile_admin(username: str, request: Request):
    """Retourne le profil d'un utilisateur (admin uniquement)."""
    from .auth import get_user_profile, is_admin
    if not is_admin(request):
        raise HTTPException(403, "Accès réservé aux administrateurs")
    profile = get_user_profile(username)
    if not profile:
        raise HTTPException(404, "Utilisateur non trouvé")
    return profile


@config_router.put("/auth/users/{username}")
def update_user_admin(
    username:     str,
    request:      Request,
    role:         str | None = Query(None),
    display_name: str | None = Query(None),
    email:        str | None = Query(None),
    phone:        str | None = Query(None),
    password:     str | None = Query(None),
):
    """Met à jour le rôle et le profil d'un utilisateur (admin uniquement)."""
    from .auth import admin_update_user, is_admin
    if not is_admin(request):
        raise HTTPException(403, "Accès réservé aux administrateurs")
    result = admin_update_user(username, role=role, display_name=display_name,
                               email=email, phone=phone, password=password or None)
    if not result:
        raise HTTPException(404, "Utilisateur non trouvé")
    return result


@config_router.delete("/auth/users/{username}")
def delete_user(username: str, request: Request):
    """Supprime un utilisateur."""
    from .auth import remove_user, is_admin
    if not is_admin(request):
        raise HTTPException(403, "Accès réservé aux administrateurs")
    return remove_user(username)


@config_router.get("/zones")
def get_zones():
    """Retourne la configuration des zones géographiques par capteur."""
    from ..alerts.zones_config import load_zones
    return load_zones()


@config_router.post("/zones/{sensor}")
def update_zone(
    sensor: str,
    zone:  str   = Query(None, description="Nom de la zone"),
    label: str   = Query(None, description="Label du capteur"),
    lat:   float = Query(None, description="Latitude"),
    lon:   float = Query(None, description="Longitude"),
):
    """Met à jour la zone géographique d'un capteur."""
    from ..alerts.zones_config import update_sensor_zone
    if sensor not in ("temperature", "turbidity", "ph"):
        raise HTTPException(400, "Capteur invalide")
    return update_sensor_zone(sensor=sensor, zone=zone, label=label, lat=lat, lon=lon)


@config_router.get("/telegram")
def get_telegram_config():
    """Retourne la configuration Telegram (token masqué)."""
    from ..alerts.telegram_notifier import get_config
    return get_config()


@config_router.post("/telegram")
def set_telegram_config(
    bot_token: str  = Query(""),
    chat_id:   str  = Query(""),
    active:    bool = Query(False),
    min_level: str  = Query("CRITICAL", description="CRITICAL | WARNING"),
):
    """Configure la notification Telegram."""
    from ..alerts.telegram_notifier import set_config
    if min_level not in ("CRITICAL", "WARNING"):
        raise HTTPException(400, "min_level doit être CRITICAL ou WARNING")
    return set_config(
        bot_token=bot_token or None,
        chat_id=chat_id or None,
        active=active,
        min_level=min_level,
    )


@config_router.get("/mqtt")
def get_mqtt_config():
    """Retourne la configuration MQTT (mot de passe masqué)."""
    from ..alerts.mqtt_publisher import get_config
    return get_config()


@config_router.post("/mqtt")
def set_mqtt_config(
    broker:       str  = Query("localhost"),
    port:         int  = Query(1883),
    username:     str  = Query(""),
    password:     str  = Query(""),
    active:       bool = Query(False),
    topic_prefix: str  = Query("iot"),
):
    """Configure le broker MQTT."""
    from ..alerts.mqtt_publisher import set_config
    return set_config(
        broker=broker, port=port,
        username=username or None,
        password=password if password != "***" else None,
        active=active,
        topic_prefix=topic_prefix,
    )


@config_router.get("/thresholds/time-rules")
def get_time_rules(sensor: str = Query(..., description="temperature | turbidity | ph")):
    """Retourne les règles horaires pour un capteur."""
    from ..alerts.threshold_config import get_time_rules
    if sensor not in ("temperature", "turbidity", "ph"):
        raise HTTPException(400, "Capteur invalide")
    return {"sensor": sensor, "time_rules": get_time_rules(sensor)}


@config_router.post("/thresholds/time-rules")
def set_time_rules(
    sensor: str = Query(..., description="temperature | turbidity | ph"),
    rules:  str = Query("[]", description="JSON array de règles horaires"),
):
    """Met à jour les règles horaires pour un capteur."""
    import json as _json
    from ..alerts.threshold_config import update_time_rules
    if sensor not in ("temperature", "turbidity", "ph"):
        raise HTTPException(400, "Capteur invalide")
    try:
        time_rules = _json.loads(rules)
        if not isinstance(time_rules, list):
            raise ValueError("tableau attendu")
    except Exception as e:
        raise HTTPException(400, f"JSON invalide : {e}")
    result = update_time_rules(sensor, time_rules)
    return {"sensor": sensor, "time_rules": result.get("time_rules", [])}


@config_router.get("/escalation")
def get_escalation_config():
    """Retourne la configuration d'escalade automatique."""
    import json, os
    path = os.path.join(os.getcwd(), "escalation_config.json")
    try:
        with open(path) as f:
            return json.load(f)
    except Exception:
        return {"active": True, "delay_minutes": 10, "notify_telegram": True,
                "notify_email": False, "notify_webhook": True}


@config_router.post("/escalation")
def set_escalation_config(
    active:           bool = Query(True),
    delay_minutes:    int  = Query(10, ge=1, le=1440),
    notify_telegram:  bool = Query(True),
    notify_email:     bool = Query(False),
    notify_webhook:   bool = Query(True),
):
    """Met à jour la configuration d'escalade automatique."""
    import json, os
    cfg = {
        "active": active,
        "delay_minutes": delay_minutes,
        "notify_telegram": notify_telegram,
        "notify_email": notify_email,
        "notify_webhook": notify_webhook,
    }
    path = os.path.join(os.getcwd(), "escalation_config.json")
    try:
        with open(path, "w") as f:
            json.dump(cfg, f, indent=2)
    except Exception as e:
        raise HTTPException(500, f"Erreur sauvegarde : {e}")
    return cfg


@config_router.post("/email")
def set_email_config(
    smtp_host: str  = Query("smtp.gmail.com"),
    smtp_port: int  = Query(587),
    username:  str  = Query(""),
    password:  str  = Query(""),
    from_addr: str  = Query(""),
    to_addr:   str  = Query(""),
    active:    bool = Query(False),
):
    """Configure la notification email SMTP."""
    from ..alerts.email_notifier import set_config
    return set_config(
        smtp_host=smtp_host, smtp_port=smtp_port,
        username=username, password=password if password != "***" else None,
        from_addr=from_addr, to_addr=to_addr, active=active,
    )


# ── /api/config/device-keys ─────────────────────────────────────────────────

@config_router.get("/device-keys")
def list_device_keys(db: Session = Depends(get_db)):
    """Liste toutes les clés API appareils."""
    from ..alerts.models import DeviceKey
    keys = db.query(DeviceKey).order_by(DeviceKey.created_at.desc()).all()
    return [k.to_dict() for k in keys]


@config_router.post("/device-keys")
def create_device_key(
    name:      str = Query(..., description="Nom de la clé"),
    device_id: str = Query("", description="Identifiant appareil"),
    _: None = Depends(require_write),
    db: Session = Depends(get_db),
):
    """Crée une nouvelle clé API pour un appareil."""
    import secrets, hashlib
    from ..alerts.models import DeviceKey
    raw = secrets.token_urlsafe(32)
    key = DeviceKey(
        name=name,
        key_hash=hashlib.sha256(raw.encode()).hexdigest(),
        device_id=device_id or None,
        active=True,
    )
    db.add(key)
    db.commit()
    db.refresh(key)
    d = key.to_dict()
    d["key"] = raw  # affiché une seule fois
    return d


@config_router.delete("/device-keys/{key_id}")
def delete_device_key(
    key_id: int,
    _: None = Depends(require_write),
    db: Session = Depends(get_db),
):
    """Supprime une clé API appareil."""
    from ..alerts.models import DeviceKey
    k = db.query(DeviceKey).filter(DeviceKey.id == key_id).first()
    if not k:
        raise HTTPException(status_code=404, detail="Clé non trouvée")
    db.delete(k)
    db.commit()
    return {"message": "Clé supprimée"}


# ── /api/config/backup ───────────────────────────────────────────────────────

@config_router.post("/backup/now")
def backup_now(_: None = Depends(require_write)):
    """Déclenche une sauvegarde immédiate de la base SQLite."""
    from ..alerts.backup import run_backup
    path = run_backup()
    if path is None:
        raise HTTPException(status_code=500, detail="Backup échoué")
    return {"message": "Backup créé", "path": path}


@config_router.get("/backup/list")
def list_backups_endpoint():
    """Liste les sauvegardes disponibles."""
    from ..alerts.backup import list_backups
    return list_backups()


# ── /api/config/thresholds/recalibrate ──────────────────────────────────────

@config_router.post("/thresholds/recalibrate")
def recalibrate_now(
    window_hours: int = Query(24, ge=1, le=720),
    _: None = Depends(require_write),
    db: Session = Depends(get_db),
):
    """Lance la recalibration automatique des seuils."""
    from ..alerts.recalibration import recalibrate_thresholds
    changes = recalibrate_thresholds(db, window_hours=window_hours)
    return {"message": "Recalibration terminée", "changes": changes}


# ── /api/config/sensors ──────────────────────────────────────────────────────

@config_router.get("/sensors")
def get_sensors_config():
    """Retourne la configuration des capteurs (sensors_config.json)."""
    import json
    from pathlib import Path
    cfg_file = Path(__file__).parent.parent.parent / "sensors_config.json"
    if cfg_file.exists():
        return json.loads(cfg_file.read_text(encoding="utf-8"))
    # Fallback vers SENSOR_CONFIG
    try:
        from ..simulator.config import SENSOR_CONFIG
        return SENSOR_CONFIG
    except Exception:
        return {}


@config_router.post("/sensors/{sensor_name}")
def update_sensor_config(
    sensor_name: str,
    warning_min:  float | None = Query(None),
    warning_max:  float | None = Query(None),
    critical_min: float | None = Query(None),
    critical_max: float | None = Query(None),
    unit:         str   | None = Query(None),
    _: None = Depends(require_write),
):
    """Met à jour la configuration d'un capteur dans sensors_config.json."""
    import json
    from pathlib import Path
    cfg_file = Path(__file__).parent.parent.parent / "sensors_config.json"
    try:
        from ..simulator.config import SENSOR_CONFIG
        cfg = json.loads(cfg_file.read_text(encoding="utf-8")) if cfg_file.exists() else dict(SENSOR_CONFIG)
    except Exception:
        cfg = {}
    if sensor_name not in cfg:
        raise HTTPException(status_code=404, detail=f"Capteur '{sensor_name}' inconnu")
    if warning_min  is not None: cfg[sensor_name]["warning_min"]  = warning_min
    if warning_max  is not None: cfg[sensor_name]["warning_max"]  = warning_max
    if critical_min is not None: cfg[sensor_name]["critical_min"] = critical_min
    if critical_max is not None: cfg[sensor_name]["critical_max"] = critical_max
    if unit         is not None: cfg[sensor_name]["unit"]         = unit
    cfg_file.write_text(json.dumps(cfg, indent=2, ensure_ascii=False), encoding="utf-8")
    return cfg[sensor_name]


# ── /api/predict/history ────────────────────────────────────────────────────

@config_router.get("/predict/history")
def get_predict_history(
    sensor: str | None = Query(None),
    limit:  int        = Query(10, ge=1, le=100),
    db: Session = Depends(get_db),
):
    """Retourne l'historique des prédictions."""
    from ..alerts.models import PredictionRecord
    q = db.query(PredictionRecord)
    if sensor:
        q = q.filter(PredictionRecord.sensor == sensor)
    records = q.order_by(PredictionRecord.created_at.desc()).limit(limit).all()
    return [r.to_dict() for r in records]


# ── /api/lang ────────────────────────────────────────────────────────────────

@config_router.post("/lang")
def set_lang(lang: str = Query("fr", pattern="^(fr|en)$")):
    """Change la langue de l'interface (cookie lang)."""
    from fastapi.responses import JSONResponse
    resp = JSONResponse({"lang": lang})
    resp.set_cookie("lang", lang, max_age=86400 * 365, httponly=False, samesite="lax")
    return resp
